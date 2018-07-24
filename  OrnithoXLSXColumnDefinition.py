# -*- coding: utf-8 -*-
#
# http://zetcode.com/articles/openpyxl/
#
import sys
import os
from openpyxl import *
from openpyxl.utils import *

__version__ = "0.1.0alpha1"


class OrnithoXLSXColumnDefinition:
    """Class to handle XLSX export data from Ornitho.de"""

    def __init__(self, pathXLSX):
        self.parseFile(pathXLSX)
        return

    def parseFile(self, pathXLSX):
        """Parse XLSX-File to extract the structure of the table (cell coordinates, column names and data types)"""

        if not os.path.exists(pathXLSX):
            raise FileNotFoundError(
                "Eingabepfad <"+pathXLSX+"> existiert nicht")
        if not os.path.isfile(pathXLSX):
            raise FileNotFoundError(
                "Eingabepfad <"+pathXLSX+"> ist keine Datei")

        book = load_workbook(pathXLSX)
        self.sheet = book["Export"]
        row_count = self.sheet.max_row
        column_count = self.sheet.max_column

        self.columns = {}
        # Parsen des Sheets um die Zuordnung Spaltennummer - Spaltennamen - Datentyp zu ermitteln.
        for col in self.sheet.iter_cols(min_row=1, max_row=1, max_col=column_count):
            for cell in col:
                name = cell.internal_value
                col = coordinate_from_string(cell.coordinate)[0]
                # Hole die Beschreibung aus der 2.ten Reihe der aktuellen Spalte
                descr = self.sheet[col + "2"].internal_value
                # Hole einen Beispielwert aus der 3.ten Reihe der aktuellen Spalte
                typ = self.sheet[col + "3"].internal_value.__class__.__name__
                self.columns[name] = (col, typ, descr)


test = OrnithoXLSXColumnDefinition(
    "w:\Develop\OrnithoXLSXImport\data\export_20180514_145320.xlsx")
print(test.columns)
